"""
SDS情報DB 正規化スクリプト v2
横長・多段ヘッダーのDBを11テーブル構成に変換する

変更点（v1からの差分）:
- GHS_BOX1〜9 削除
- 子テーブルの 材料名・連番 削除
- chemicals → chemical_master + material_chemicals（許容濃度は非nullを優先）
- hazards   → hazard_master  + material_hazards（有害性文言を重複排除）
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import sys
import io

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

INPUT_FILE  = 'data/SDS情報DB_元データ.xlsx'
OUTPUT_FILE = 'data/SDS情報DB_正規化.xlsx'

# ---------------------------------------------------------------------------
# 1. 生データ読み込み
# ---------------------------------------------------------------------------
print("読み込み中...")
df_raw = pd.read_excel(INPUT_FILE, header=None)
data_rows = df_raw.iloc[6:].reset_index(drop=True)  # row6以降がデータ

def get(row, col):
    v = row.iloc[col] if col < len(row) else None
    if pd.isna(v) or str(v).strip() in ('nan', ''):
        return None
    return str(v).strip()

def get_num(row, col):
    v = row.iloc[col] if col < len(row) else None
    if pd.isna(v):
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None

def get_flag(row, col):
    """GHSフラグ列を 1/0 に正規化する。
    空セル・✕系 → 0、それ以外（○系）→ 1。
    ○の文字コード違い（U+25CB/U+25EF等）に依存しないよう否定側で判定する。
    """
    v = get(row, col)
    if v is None:
        return 0
    if v.lower() in ('✕', '✗', '×', 'x', '0', '0.0', 'false', 'no', '-'):
        return 0
    return 1

# ---------------------------------------------------------------------------
# 2. 定数定義
# ---------------------------------------------------------------------------
FIRE_AGENT_COLS = {
    198: '霧状水',
    199: '棒状水',
    200: '炭酸ガス',
    201: '泡',
    202: 'リン酸塩粉末',
    203: '炭酸塩粉末',
    204: '乾燥砂',
    205: '霧状強化液',
    206: '棒状強化液',
    207: 'その他',
}

FIRST_AID_CATS = {
    178: ('吸入',           [178, 179, 180, 181]),
    182: ('皮膚付着',       [182, 183, 184, 185]),
    186: ('眼',             [186, 187, 188, 189]),
    190: ('飲み込んだ場合', [190, 191, 192, 193]),
    194: ('応急措置者の保護',[194, 195, 196, 197]),
}

PROTECTION_COLS = {
    170: '皮膚',
    172: '眼',
    174: '呼吸',
    176: '手',
}

SPILL_CATS = {
    '人体に関する': [216, 217, 218, 219],
    '環境に関する': [220, 221, 222, 223],
    '封じ込め方法': [224, 225, 226, 227],
}

# ---------------------------------------------------------------------------
# 3. テーブル構築
# ---------------------------------------------------------------------------
materials         = []
material_chemicals = []
material_hazards  = []
risk_reductions   = []
protective_equip  = []
first_aids        = []
fire_agents       = []
fire_methods      = []
spill_responses   = []

# chemical_master: 化学物質名 → {chemical_id, 許容濃度}
chemical_name_to_id = {}
chemical_id_to_row  = {}
chemical_id_seq     = 0

# hazard_master: 有害性内容 → hazard_id
hazard_text_to_id   = {}
hazard_id_to_row    = {}
hazard_id_seq       = 0

# 許容濃度の競合ログ
kyoyo_conflicts = []

material_id = 0

for _, row in data_rows.iterrows():
    mat_name = get(row, 2)
    if not mat_name:
        continue

    material_id += 1
    mid = material_id

    # ---- 材料マスタ ----
    materials.append({
        'material_id':    mid,
        '頭文字':         get(row, 1),
        '材料名':         mat_name,
        'GHS_可燃':       get_flag(row, 3),
        'GHS_支燃':       get_flag(row, 4),
        'GHS_爆発':       get_flag(row, 5),
        'GHS_腐食':       get_flag(row, 6),
        'GHS_ガス':       get_flag(row, 7),
        'GHS_毒性1':      get_flag(row, 8),
        'GHS_毒性2':      get_flag(row, 9),
        'GHS_環境':       get_flag(row, 10),
        'GHS_臓器':       get_flag(row, 11),
        '見積もったリスク': get_num(row, 61),
        '要変更':         get(row, 228),
        '特化則':         get(row, 229),
        '有機則':         get(row, 230),
        '鉛中毒':         get(row, 231),
        '記録保存':       get(row, 232),
    })

    # ---- 有害性 → hazard_master + material_hazards ----
    for i in range(20):
        content_col = 21 + i * 2
        content = get(row, content_col)
        score   = get_num(row, content_col + 1)
        if not content:
            continue

        if content not in hazard_text_to_id:
            hazard_id_seq += 1
            hazard_text_to_id[content] = hazard_id_seq
            hazard_id_to_row[hazard_id_seq] = {'hazard_id': hazard_id_seq, '有害性内容': content}

        material_hazards.append({
            'material_id': mid,
            'hazard_id':   hazard_text_to_id[content],
            '点数':        score,
        })

    # ---- リスク低減措置 ----
    for i, label in enumerate('ABCDEFGHIJKL'):
        content = get(row, 62 + i)
        if content:
            risk_reductions.append({'material_id': mid, '記号': label, '内容': content})

    # ---- 含有化学物質 → chemical_master + material_chemicals ----
    for i in range(16):  # A〜P
        base  = 74 + i * 6
        name  = get(row, base)
        if not name:
            continue

        kyoyo = get(row, base + 5)

        if name not in chemical_name_to_id:
            chemical_id_seq += 1
            chemical_name_to_id[name] = chemical_id_seq
            chemical_id_to_row[chemical_id_seq] = {
                'chemical_id': chemical_id_seq,
                '化学物質名':  name,
                '許容濃度':    kyoyo,
            }
        else:
            cid = chemical_name_to_id[name]
            existing = chemical_id_to_row[cid]
            if existing['許容濃度'] is None and kyoyo is not None:
                existing['許容濃度'] = kyoyo  # 非nullで上書き
            elif (existing['許容濃度'] is not None
                  and kyoyo is not None
                  and existing['許容濃度'] != kyoyo):
                kyoyo_conflicts.append((name, existing['許容濃度'], kyoyo))

        material_chemicals.append({
            'material_id':    mid,
            'chemical_id':    chemical_name_to_id[name],
            '含有率最小':     get_num(row, base + 1),
            '含有率最大':     get_num(row, base + 2),
            '推定濃度_長時間': get_num(row, base + 3),
            '推定濃度_短時間': get_num(row, base + 4),
        })

    # ---- 保護具 ----
    for col, part in PROTECTION_COLS.items():
        content = get(row, col)
        if content:
            protective_equip.append({'material_id': mid, '保護部位': part, '内容': content})

    # ---- 応急処置 ----
    for _, (cat, cols) in FIRST_AID_CATS.items():
        for col in cols:
            content = get(row, col)
            if content:
                first_aids.append({'material_id': mid, 'カテゴリ': cat, '内容': content})

    # ---- 緊急対応：消火剤 ----
    for col, agent in FIRE_AGENT_COLS.items():
        val = get(row, col)
        if val:
            fire_agents.append({'material_id': mid, '消火剤種別': agent, '適否': val})

    # ---- 緊急対応：消火方法 ----
    for col in range(208, 216):
        content = get(row, col)
        if content:
            fire_methods.append({'material_id': mid, '内容': content})

    # ---- 緊急対応：漏出時措置 ----
    for cat, cols in SPILL_CATS.items():
        for col in cols:
            content = get(row, col)
            if content:
                spill_responses.append({'material_id': mid, 'カテゴリ': cat, '内容': content})

# ---------------------------------------------------------------------------
# 4. 競合レポート
# ---------------------------------------------------------------------------
if kyoyo_conflicts:
    print(f"\n⚠️  許容濃度の競合 ({len(kyoyo_conflicts)}件) — 先に登録された値を採用:")
    seen = set()
    for name, v1, v2 in kyoyo_conflicts:
        if name not in seen:
            print(f"  {name}: '{v1}' vs '{v2}'")
            seen.add(name)
else:
    print("\n✅ 許容濃度の競合なし")

# ---------------------------------------------------------------------------
# 5. DataFrameに変換・件数表示
# ---------------------------------------------------------------------------
df_materials         = pd.DataFrame(materials)
df_chemical_master   = pd.DataFrame(chemical_id_to_row.values())
df_material_chemicals = pd.DataFrame(material_chemicals)
df_hazard_master     = pd.DataFrame(hazard_id_to_row.values())
df_material_hazards  = pd.DataFrame(material_hazards)
df_risk              = pd.DataFrame(risk_reductions)
df_protect           = pd.DataFrame(protective_equip)
df_firstaid          = pd.DataFrame(first_aids)
df_fire_agents       = pd.DataFrame(fire_agents)
df_fire_methods      = pd.DataFrame(fire_methods)
df_spill             = pd.DataFrame(spill_responses)

print("\n--- テーブル件数 ---")
print(f"材料マスタ:          {len(df_materials):>6}件")
print(f"化学物質マスタ:      {len(df_chemical_master):>6}件  (ユニーク)")
print(f"材料×化学物質:      {len(df_material_chemicals):>6}件")
print(f"有害性マスタ:        {len(df_hazard_master):>6}件  (ユニーク)")
print(f"材料×有害性:        {len(df_material_hazards):>6}件")
print(f"リスク低減措置:      {len(df_risk):>6}件")
print(f"保護具:              {len(df_protect):>6}件")
print(f"応急処置:            {len(df_firstaid):>6}件")
print(f"緊急対応_消火剤:     {len(df_fire_agents):>6}件")
print(f"緊急対応_消火方法:   {len(df_fire_methods):>6}件")
print(f"緊急対応_漏出時措置: {len(df_spill):>6}件")

# ---------------------------------------------------------------------------
# 6. Excelに書き出し
# ---------------------------------------------------------------------------
print("\n書き出し中...")

HEADER_COLOR = "4472C4"

def style_sheet(ws, df):
    header_font  = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    header_fill  = PatternFill('solid', start_color=HEADER_COLOR)
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    data_font    = Font(name='Arial', size=10)
    thin         = Side(style='thin', color='CCCCCC')
    border       = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font, cell.fill, cell.alignment, cell.border = header_font, header_fill, header_align, border

    for row_idx, row in enumerate(df.itertuples(index=False), 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font   = data_font
            cell.border = border
            cell.alignment = Alignment(vertical='top', wrap_text=False)

    for col_idx, col_name in enumerate(df.columns, 1):
        max_len = len(str(col_name))
        for row_idx in range(2, min(len(df) + 2, 52)):
            v = ws.cell(row=row_idx, column=col_idx).value
            if v:
                max_len = max(max_len, max(len(ln) for ln in str(v).split('\n')))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len * 1.2 + 2, 8), 50)

    ws.freeze_panes   = 'B2'
    ws.row_dimensions[1].height = 25

wb = openpyxl.Workbook()
wb.remove(wb.active)

sheets = [
    ('材料マスタ',          df_materials),
    ('化学物質マスタ',      df_chemical_master),
    ('材料×化学物質',      df_material_chemicals),
    ('有害性マスタ',        df_hazard_master),
    ('材料×有害性',        df_material_hazards),
    ('リスク低減措置',      df_risk),
    ('保護具',              df_protect),
    ('応急処置',            df_firstaid),
    ('緊急対応_消火剤',     df_fire_agents),
    ('緊急対応_消火方法',   df_fire_methods),
    ('緊急対応_漏出時措置', df_spill),
]

for sheet_name, df in sheets:
    ws = wb.create_sheet(sheet_name)
    style_sheet(ws, df)
    print(f"  → {sheet_name} ({len(df)}行)")

wb.save(OUTPUT_FILE)
print(f"\n完了: {OUTPUT_FILE}")
